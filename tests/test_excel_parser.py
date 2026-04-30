import pytest
import openpyxl
from decimal import Decimal
from datetime import date
from parsers.excel_parser import parse_excel


def test_parse_excel_basico(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Data", "Histórico", "Valor", "Saldo"])
    ws.append(["15/01/2024", "PIX RECEBIDO", "1000,00", "5000,00"])
    ws.append(["16/01/2024", "COMPRA", "-150,00", "4850,00"])
    path = str(tmp_path / "extrato.xlsx")
    wb.save(path)
    t = parse_excel(path)
    assert len(t) == 2
    assert t[0].valor == Decimal("1000.00")
    assert t[0].tipo == "Crédito"
    assert t[1].valor == Decimal("-150.00")
    assert t[1].tipo == "Débito"


def test_parse_excel_datas(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Data", "Histórico", "Valor"])
    ws.append(["2024-01-15", "PIX", "500,00"])
    path = str(tmp_path / "extrato.xlsx")
    wb.save(path)
    t = parse_excel(path)
    assert t[0].data == date(2024, 1, 15)


def test_parse_excel_colunas_credito_debito(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Data", "Histórico", "Crédito", "Débito", "Saldo"])
    ws.append(["15/01/2024", "PIX", "1000,00", "", "5000,00"])
    ws.append(["16/01/2024", "BOLETO", "", "200,00", "4800,00"])
    path = str(tmp_path / "extrato.xlsx")
    wb.save(path)
    t = parse_excel(path)
    assert len(t) == 2
    assert t[0].valor == Decimal("1000.00")
    assert t[1].valor == Decimal("-200.00")
