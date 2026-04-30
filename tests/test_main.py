import os
import pytest
import openpyxl
from main import processar_extrato

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")


def test_processar_extrato_ofx(tmp_path):
    output = str(tmp_path / "relatorio.xlsx")
    resultado = processar_extrato(os.path.join(FIXTURES, "sample.ofx"), output)
    assert os.path.exists(resultado)
    wb = openpyxl.load_workbook(resultado)
    assert "Extrato" in wb.sheetnames
    assert "Resumo" in wb.sheetnames


def test_processar_extrato_csv_nubank(tmp_path):
    output = str(tmp_path / "relatorio.xlsx")
    resultado = processar_extrato(os.path.join(FIXTURES, "nubank.csv"), output)
    assert os.path.exists(resultado)
    wb = openpyxl.load_workbook(resultado)
    ws = wb["Extrato"]
    assert ws.max_row >= 3  # cabecalho + 2 transacoes + total


def test_processar_extrato_formato_invalido(tmp_path):
    f = tmp_path / "arquivo.zip"
    f.write_bytes(b"PK")
    with pytest.raises(ValueError):
        processar_extrato(str(f), str(tmp_path / "out.xlsx"))
