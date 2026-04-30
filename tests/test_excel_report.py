import os
import pytest
import openpyxl
from decimal import Decimal
from datetime import date
from models.transaction import Transaction
from report.excel_report import gerar_relatorio

TRANSACOES = [
    Transaction(date(2024, 1, 15), "PIX RECEBIDO - JOAO",  Decimal("1000.00"),  "Crédito", Decimal("5000.00")),
    Transaction(date(2024, 1, 16), "COMPRA SUPERMERCADO",   Decimal("-150.00"),  "Débito",  Decimal("4850.00")),
    Transaction(date(2024, 1, 17), "SALARIO",               Decimal("3000.00"),  "Crédito", Decimal("7850.00")),
    Transaction(date(2024, 1, 20), "ALUGUEL",               Decimal("-1200.00"), "Débito",  Decimal("6650.00")),
]


def test_gerar_relatorio_cria_arquivo(tmp_path):
    output = str(tmp_path / "relatorio.xlsx")
    gerar_relatorio(TRANSACOES, output)
    assert os.path.exists(output)


def test_relatorio_tem_duas_abas(tmp_path):
    output = str(tmp_path / "relatorio.xlsx")
    gerar_relatorio(TRANSACOES, output)
    wb = openpyxl.load_workbook(output)
    assert "Extrato" in wb.sheetnames
    assert "Resumo" in wb.sheetnames


def test_extrato_tem_cabecalho_correto(tmp_path):
    output = str(tmp_path / "relatorio.xlsx")
    gerar_relatorio(TRANSACOES, output)
    wb = openpyxl.load_workbook(output)
    ws = wb["Extrato"]
    cabecalho = [ws.cell(1, c).value for c in range(1, 6)]
    assert cabecalho[0] == "Data"
    assert cabecalho[2] is not None  # Debitos
    assert cabecalho[3] is not None  # Creditos


def test_extrato_tem_linhas_corretas(tmp_path):
    output = str(tmp_path / "relatorio.xlsx")
    gerar_relatorio(TRANSACOES, output)
    wb = openpyxl.load_workbook(output)
    ws = wb["Extrato"]
    # 1 cabecalho + 4 transacoes + 1 total = 6
    assert ws.max_row == 6


def test_celulas_debito_tem_fonte_vermelha(tmp_path):
    output = str(tmp_path / "relatorio.xlsx")
    gerar_relatorio(TRANSACOES, output)
    wb = openpyxl.load_workbook(output)
    ws = wb["Extrato"]
    # Linha 3 = segundo lancamento (debito)
    cor = ws.cell(3, 4).font.color.rgb  # col 4 = Debitos (col 3 agora eh Categoria)
    assert cor.endswith("0000")  # termina com 0000 = vermelho (FF0000 ou C00000)


def test_resumo_tem_total_creditos(tmp_path):
    output = str(tmp_path / "relatorio.xlsx")
    gerar_relatorio(TRANSACOES, output)
    wb = openpyxl.load_workbook(output)
    ws = wb["Resumo"]
    valores = [ws.cell(r, 2).value for r in range(1, ws.max_row + 1)]
    # 1000 + 3000 = 4000
    assert 4000.0 in valores
