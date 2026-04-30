"""
Gerador de relatorio Excel de conciliacao bancaria.
Modos: "simples" (1 aba resumo) e "detalhado" (3 abas completas).
"""
from collections import defaultdict
from decimal import Decimal
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from models.transaction import Transaction

# ── Paleta de cores ──────────────────────────────────────────────────────────
C_AZUL_ESCURO   = "1F3864"
C_AZUL_MEDIO    = "2E75B6"
C_AZUL_CLARO    = "BDD7EE"
C_BRANCO        = "FFFFFF"
C_CINZA_CLARO   = "F2F2F2"
C_VERDE_ESCURO  = "1E7145"
C_VERDE_CLARO   = "E2EFDA"
C_VERMELHO      = "C00000"
C_VERMELHO_CLARO= "FCE4D6"
C_AMARELO       = "FFD966"
C_AMARELO_CLARO = "FFF2CC"
C_LARANJA       = "ED7D31"
C_CINZA_FONTE   = "595959"

FMT_MOEDA = 'R$ #,##0.00;[RED]-R$ #,##0.00'
FMT_DATA  = 'DD/MM/YYYY'
FMT_PCT   = '0.0"%"'


def _borda(estilo="thin", cor="D9D9D9"):
    l = Side(style=estilo, color=cor)
    return Border(left=l, right=l, top=l, bottom=l)


def _borda_ext():
    """Borda externa mais marcada."""
    ext = Side(style="medium", color="2E75B6")
    thin = Side(style="thin",   color="D9D9D9")
    return Border(left=ext, right=ext, top=ext, bottom=ext)


def _fill(cor):
    return PatternFill("solid", fgColor=cor)


def _ajustar_largura(ws, minw=10, maxw=55):
    for col in ws.columns:
        maxlen = max((len(str(c.value or "")) for c in col), default=minw)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(maxlen + 3, minw), maxw)


def _titulo_celula(ws, row, col, texto, fundo=C_AZUL_ESCURO, fonte=C_BRANCO, size=11, bold=True, span=None):
    c = ws.cell(row, col, texto)
    c.fill = _fill(fundo)
    c.font = Font(bold=bold, color=fonte, size=size)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _borda("medium", C_AZUL_ESCURO)
    if span:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span)
    return c


def _kv(ws, row, col_label, col_val, label, valor, fmt=None,
         fundo_label=C_AZUL_CLARO, fundo_val=C_BRANCO,
         cor_val="000000", bold_val=False):
    lc = ws.cell(row, col_label, label)
    lc.fill = _fill(fundo_label)
    lc.font = Font(bold=True, color=C_AZUL_ESCURO, size=10)
    lc.border = _borda()
    lc.alignment = Alignment(vertical="center", wrap_text=True)
    vc = ws.cell(row, col_val, valor)
    vc.fill = _fill(fundo_val)
    vc.font = Font(bold=bold_val, color=cor_val, size=10)
    vc.border = _borda()
    vc.alignment = Alignment(horizontal="right", vertical="center")
    if fmt:
        vc.number_format = fmt
    return vc


# ─────────────────────────────────────────────────────────────────────────────
#  BANNER DE SAUDE FINANCEIRA
# ─────────────────────────────────────────────────────────────────────────────
def _banner_saude(ws, row_start, total_cred, total_deb, saldo, n_cols=5):
    saudavel = saldo >= 0
    cor_fundo = C_VERDE_CLARO if saudavel else C_VERMELHO_CLARO
    cor_fonte = C_VERDE_ESCURO if saudavel else C_VERMELHO
    emoji = "SAUDE FINANCEIRA: OK" if saudavel else "ATENCAO: GASTOS SUPERAM RECEITAS!"
    pct = (total_deb / total_cred * 100) if total_cred else Decimal("0")

    ws.row_dimensions[row_start].height = 30
    c = ws.cell(row_start, 1, emoji)
    c.fill = _fill(cor_fundo)
    c.font = Font(bold=True, color=cor_fonte, size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _borda("medium", cor_fonte)
    ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=n_cols)

    r2 = row_start + 1
    ws.row_dimensions[r2].height = 18
    msg = f"Voce gastou {float(pct):.1f}% da sua renda neste periodo"
    c2 = ws.cell(r2, 1, msg)
    c2.fill = _fill(cor_fundo)
    c2.font = Font(italic=True, color=cor_fonte, size=10)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = _borda("thin", cor_fonte)
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=n_cols)
    return row_start + 2


# ─────────────────────────────────────────────────────────────────────────────
#  ABA EXTRATO (modo detalhado)
# ─────────────────────────────────────────────────────────────────────────────
def _aba_extrato(ws, transacoes: List[Transaction]):
    from parsers.categorizer import categorizar
    ws.sheet_view.showGridLines = False

    # Cabecalho
    ws.row_dimensions[1].height = 28
    for c, (titulo, w) in enumerate([
        ("Data", 13), ("Descricao", 45), ("Categoria", 22),
        ("Debitos (R$)", 16), ("Creditos (R$)", 16), ("Saldo (R$)", 16)
    ], 1):
        _titulo_celula(ws, 1, c, titulo)
        ws.column_dimensions[get_column_letter(c)].width = w

    for i, t in enumerate(transacoes):
        row = i + 2
        fundo = C_CINZA_CLARO if i % 2 == 0 else C_BRANCO
        cat = categorizar(t)
        debito  = float(abs(t.valor)) if t.tipo == "Debito"  else None
        credito = float(t.valor)      if t.tipo == "Credito" else None
        saldo   = float(t.saldo)      if t.saldo is not None else None
        vals = [t.data, t.descricao, cat, debito, credito, saldo]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row, c, val)
            cell.fill = _fill(fundo)
            cell.border = _borda()
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 2))
            if c == 1:
                cell.number_format = FMT_DATA
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 3:
                cell.font = Font(color=C_AZUL_MEDIO, size=9, italic=True)
            elif c == 4 and val is not None:
                cell.number_format = FMT_MOEDA
                cell.font = Font(color=C_VERMELHO)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c == 5 and val is not None:
                cell.number_format = FMT_MOEDA
                cell.font = Font(color=C_VERDE_ESCURO)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c == 6 and val is not None:
                cell.number_format = FMT_MOEDA
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Linha de totais
    tr = len(transacoes) + 2
    ws.row_dimensions[tr].height = 22
    tot_deb  = sum(abs(t.valor) for t in transacoes if t.tipo == "Debito")
    tot_cred = sum(t.valor      for t in transacoes if t.tipo == "Credito")
    for c in range(1, 7):
        cell = ws.cell(tr, c)
        cell.fill = _fill(C_AZUL_ESCURO)
        cell.font = Font(bold=True, color=C_BRANCO, size=11)
        cell.border = _borda("medium", C_AZUL_ESCURO)
    ws.cell(tr, 2, "TOTAL GERAL").alignment = Alignment(horizontal="right", vertical="center")
    c4 = ws.cell(tr, 4, float(tot_deb));  c4.number_format = FMT_MOEDA; c4.alignment = Alignment(horizontal="right", vertical="center")
    c5 = ws.cell(tr, 5, float(tot_cred)); c5.number_format = FMT_MOEDA; c5.alignment = Alignment(horizontal="right", vertical="center")
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
#  ABA RESUMO (usado em ambos os modos)
# ─────────────────────────────────────────────────────────────────────────────
def _aba_resumo(ws, transacoes: List[Transaction], limites: dict):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    if not transacoes:
        ws.cell(1, 1, "Nenhuma transacao encontrada")
        return

    datas    = sorted(t.data for t in transacoes)
    tot_cred = sum(t.valor      for t in transacoes if t.tipo == "Credito")
    tot_deb  = sum(abs(t.valor) for t in transacoes if t.tipo == "Debito")
    saldo    = tot_cred - tot_deb
    qtd      = len(transacoes)

    # Titulo
    row = 1
    ws.row_dimensions[row].height = 32
    _titulo_celula(ws, row, 1, "RESUMO DA CONCILIACAO BANCARIA", size=13, span=4)
    row += 1

    # Banner de saude
    row = _banner_saude(ws, row, tot_cred, tot_deb, saldo, n_cols=4)
    row += 1  # espaco

    # Bloco de totais gerais
    ws.row_dimensions[row].height = 22
    _titulo_celula(ws, row, 1, "TOTAIS DO PERIODO", size=11, span=4)
    row += 1
    _kv(ws, row, 1, 2, "Periodo",          f"{datas[0].strftime('%d/%m/%Y')} a {datas[-1].strftime('%d/%m/%Y')}")
    row += 1
    _kv(ws, row, 1, 2, "Qtd. de lancamentos", qtd)
    row += 1
    _kv(ws, row, 1, 2, "Total de Entradas",  float(tot_cred), FMT_MOEDA,
        cor_val=C_VERDE_ESCURO, bold_val=True)
    row += 1
    _kv(ws, row, 1, 2, "Total de Saidas",    float(tot_deb),  FMT_MOEDA,
        cor_val=C_VERMELHO, bold_val=True)
    row += 1
    cor_saldo = C_VERDE_ESCURO if saldo >= 0 else C_VERMELHO
    fundo_saldo = C_VERDE_CLARO if saldo >= 0 else C_VERMELHO_CLARO
    _kv(ws, row, 1, 2, "Saldo Liquido",     float(saldo),    FMT_MOEDA,
        fundo_val=fundo_saldo, cor_val=cor_saldo, bold_val=True)
    row += 2  # espaco

    # Analise por categoria de DEBITO
    ws.row_dimensions[row].height = 22
    _titulo_celula(ws, row, 1, "ANALISE DE GASTOS POR CATEGORIA", size=11, span=4)
    row += 1

    # Cabecalho da tabela de categorias
    for c, label in enumerate(["Categoria", "Total (R$)", "% da Renda", "Status"], 1):
        _titulo_celula(ws, row, c, label, fundo=C_AZUL_MEDIO, size=10)
    row += 1

    from parsers.categorizer import agrupar_por_categoria
    grupos = agrupar_por_categoria(transacoes)
    grupos_deb = {k: v for k, v in grupos.items() if v["tipo"] == "Debito"}
    grupos_deb_sorted = sorted(grupos_deb.items(), key=lambda x: x[1]["total"], reverse=True)

    for cat, dados in grupos_deb_sorted:
        total_cat = dados["total"]
        pct = float(total_cat / tot_cred * 100) if tot_cred else 0.0
        limite = limites.get(cat, limites.get("Outros Debitos", 15))
        excede = pct > limite

        fundo = C_AMARELO_CLARO if excede else C_BRANCO
        cor_status = C_LARANJA if excede else C_VERDE_ESCURO
        status = f"ALTO (limite: {limite:.0f}%)" if excede else "OK"

        ws.cell(row, 1, cat).fill   = _fill(fundo)
        ws.cell(row, 1).font        = Font(color=C_AZUL_ESCURO)
        ws.cell(row, 1).border      = _borda()
        ws.cell(row, 1).alignment   = Alignment(vertical="center")

        c2 = ws.cell(row, 2, float(total_cat))
        c2.number_format = FMT_MOEDA
        c2.fill  = _fill(fundo)
        c2.font  = Font(color=C_VERMELHO, bold=True)
        c2.border = _borda()
        c2.alignment = Alignment(horizontal="right", vertical="center")

        c3 = ws.cell(row, 3, round(pct, 1))
        c3.number_format = '0.0"%"'
        c3.fill  = _fill(fundo)
        c3.font  = Font(color=C_CINZA_FONTE, bold=excede)
        c3.border = _borda()
        c3.alignment = Alignment(horizontal="right", vertical="center")

        c4 = ws.cell(row, 4, status)
        c4.fill  = _fill(C_AMARELO_CLARO if excede else C_VERDE_CLARO)
        c4.font  = Font(bold=True, color=cor_status)
        c4.border = _borda()
        c4.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    row += 1  # espaco

    # Tabela mensal se periodo > 1 mes
    meses = defaultdict(lambda: {"cred": Decimal("0"), "deb": Decimal("0")})
    for t in transacoes:
        k = (t.data.year, t.data.month)
        if t.tipo == "Credito":
            meses[k]["cred"] += t.valor
        else:
            meses[k]["deb"] += abs(t.valor)

    if len(meses) > 1:
        ws.row_dimensions[row].height = 22
        _titulo_celula(ws, row, 1, "EVOLUCAO MENSAL", size=11, span=4)
        row += 1
        for c, label in enumerate(["Mes/Ano", "Entradas (R$)", "Saidas (R$)", "Saldo (R$)"], 1):
            _titulo_celula(ws, row, c, label, fundo=C_AZUL_MEDIO, size=10)
        row += 1
        for (ano, mes), v in sorted(meses.items()):
            import calendar
            nome_mes = calendar.month_abbr[mes].upper()
            sal_mes  = v["cred"] - v["deb"]
            fundo_mes = C_VERDE_CLARO if sal_mes >= 0 else C_VERMELHO_CLARO
            cor_sal   = C_VERDE_ESCURO if sal_mes >= 0 else C_VERMELHO
            for c, (val, fmt, cor) in enumerate([
                (f"{nome_mes}/{ano}", None, C_AZUL_ESCURO),
                (float(v["cred"]),  FMT_MOEDA, C_VERDE_ESCURO),
                (float(v["deb"]),   FMT_MOEDA, C_VERMELHO),
                (float(sal_mes),    FMT_MOEDA, cor_sal),
            ], 1):
                cell = ws.cell(row, c, val)
                cell.fill  = _fill(fundo_mes if c == 4 else C_BRANCO)
                cell.font  = Font(color=cor, bold=(c == 4))
                cell.border = _borda()
                cell.alignment = Alignment(horizontal="right" if c > 1 else "center", vertical="center")
                if fmt:
                    cell.number_format = fmt
            row += 1


# ─────────────────────────────────────────────────────────────────────────────
#  ABA SIMPLES (modo simples = apenas resumo executivo)
# ─────────────────────────────────────────────────────────────────────────────
def _aba_simples(ws, transacoes: List[Transaction], limites: dict):
    """Visao compacta: saude financeira + totais + top categorias. Sem lista de transacoes."""
    _aba_resumo(ws, transacoes, limites)
    ws.title = "Resumo Simples"


# ─────────────────────────────────────────────────────────────────────────────
#  FUNCAO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────
def gerar_relatorio(transacoes: List[Transaction], output_path: str,
                    modo: str = "detalhado") -> str:
    """
    Gera relatorio Excel.
    modo='simples'  -> 1 aba: Resumo executivo + saude + categorias
    modo='detalhado' -> 3 abas: Extrato completo + Resumo + (mensal se >1 mes)
    """
    import sys, os
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from config.limites_parser import carregar_limites

    # Normaliza tipo sem acento para comparacao interna
    for t in transacoes:
        if t.tipo not in ("Debito", "Credito"):
            t.tipo = "Credito" if t.valor >= 0 else "Debito"

    limites = carregar_limites()
    wb = openpyxl.Workbook()

    if modo == "simples":
        ws = wb.active
        ws.title = "Resumo"
        _aba_simples(ws, transacoes, limites)
    else:
        ws_ext = wb.active
        ws_ext.title = "Extrato"
        _aba_extrato(ws_ext, transacoes)

        ws_res = wb.create_sheet("Resumo")
        _aba_resumo(ws_res, transacoes, limites)

    wb.save(output_path)
    return output_path
